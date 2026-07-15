"""Excel export for GitHub repositories."""

from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_EXCEL_FILE = "github_repos.xlsx"

# Characters Excel/LibreOffice treat as the start of a formula. Values that
# begin with one are prefixed with an apostrophe so they render as literal
# text — this prevents CSV/formula injection from untrusted repo metadata.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_text(value: str | None) -> str:
    text = value or ""
    if text.startswith(_FORMULA_PREFIXES):
        return "'" + text
    return text


def write_excel(repos: list[dict[str, Any]], path: str = DEFAULT_EXCEL_FILE) -> None:
    """Write repos to an .xlsx file sorted by Last Updated (descending).

    Archived repos are pushed to the bottom (still date-sorted among
    themselves) and rendered with strikethrough text. The sheet has four
    columns — Repository, Description, Visibility, Last Updated — with
    banded rows, a frozen header, an autofilter, and landscape page
    setup that fits to width when printed.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned a workbook with no active sheet")
    ws.title = "Repositories"

    headers = ["Repository", "Description", "Visibility", "Last Updated"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    repos_sorted = sorted(
        repos,
        key=lambda r: (bool(r.get("archived")), -_updated_epoch(r.get("updated_at"))),
    )
    band_fill = PatternFill("solid", fgColor="F2F2F2")
    archived_font = Font(strike=True)

    for i, repo in enumerate(repos_sorted, start=2):
        archived = bool(repo.get("archived"))
        updated_dt = _parse_updated(repo.get("updated_at"))
        ws.cell(row=i, column=1, value=_safe_text(repo.get("name")))
        ws.cell(row=i, column=2, value=_safe_text(repo.get("description")))
        ws.cell(row=i, column=3, value=_safe_text(repo.get("visibility")).capitalize())
        cell = ws.cell(row=i, column=4, value=updated_dt)
        cell.number_format = "yyyy-mm-dd hh:mm"

        if i % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=i, column=col_idx).fill = band_fill
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=col_idx)
            c.alignment = Alignment(vertical="center", wrap_text=(col_idx == 2))
            if archived:
                c.font = archived_font

    _apply_layout(ws, len(headers))
    wb.save(path)


def _updated_epoch(value: str | None) -> float:
    dt = _parse_updated(value)
    return dt.timestamp() if dt else 0.0


def _style_header_row(ws: Worksheet, ncols: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _parse_updated(value: str | None) -> datetime | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%dT%H:%M:%SZ")


def _apply_layout(ws: Worksheet, ncols: int) -> None:
    widths = {1: 32, 2: 80, 3: 14, 4: 20}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # ty: ignore[invalid-assignment]
    ws.print_options.gridLines = False
    ws.print_title_rows = "1:1"
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
