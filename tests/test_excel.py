"""Tests for the Excel export."""

from __future__ import annotations

from pathlib import Path

from conftest import make_repo
from openpyxl import load_workbook

from gman.excel import _safe_text, write_excel


def test_safe_text_escapes_formula_characters() -> None:
    assert _safe_text("=1+1") == "'=1+1"
    assert _safe_text("+cmd") == "'+cmd"
    assert _safe_text("-2") == "'-2"
    assert _safe_text("@ref") == "'@ref"
    assert _safe_text("normal") == "normal"
    assert _safe_text(None) == ""


def test_write_excel_orders_and_escapes(tmp_path: Path) -> None:
    repos = [
        make_repo("old", updated_at="2025-01-01T00:00:00Z", archived=True),
        make_repo("newer", updated_at="2026-06-01T00:00:00Z"),
        make_repo("evil", description="=HYPERLINK(0)", updated_at="2026-01-01T00:00:00Z"),
    ]
    out = tmp_path / "repos.xlsx"

    write_excel(repos, str(out))

    wb = load_workbook(out)
    ws = wb["Repositories"]
    names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    # Newest active first, archived pushed to the bottom.
    assert names == ["newer", "evil", "old"]
    # Formula-like description was neutralized.
    descriptions = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "'=HYPERLINK(0)" in descriptions
    # Header present.
    assert ws.cell(row=1, column=1).value == "Repository"
